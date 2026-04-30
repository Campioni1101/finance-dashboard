"""
Gera finance_data.xlsx — planilha organizada com dados migrados do finance.xlsx.
Execute: python create_template.py
"""
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.styles.differential import DifferentialStyle
from datetime import datetime, date

from data_parser import parse_finance_data, _is_fii, CALENDAR_ORDER

# ── Paleta ────────────────────────────────────────────────────────────────────
HDR_BG   = '1e3a5f'   # azul escuro — header principal
HDR_FG   = 'FFFFFF'
SUB_BG   = '2563eb'   # azul médio — sub-header
SUB_FG   = 'FFFFFF'
TITLE_BG = '0f172a'   # quase preto — título
TITLE_FG = 'e2e8f0'
ALT_BG   = 'f1f5f9'   # cinza muito claro — linhas alternadas
WHITE    = 'FFFFFF'
GREEN_BG = 'd1fae5'   # verde claro — ganho
RED_BG   = 'fee2e2'   # vermelho claro — perda
BLUE_BG  = 'dbeafe'   # azul claro — referência
YELLOW_BG= 'fef9c3'   # amarelo — atenção
BORDER_C = 'cbd5e1'   # borda suave

MONTH_PT = {
    'jan':'Janeiro','fev':'Fevereiro','mar':'Março','abr':'Abril',
    'mai':'Maio','jun':'Junho','jul':'Julho','ago':'Agosto',
    'set':'Setembro','out':'Outubro','nov':'Novembro','dez':'Dezembro'
}


# ── Helpers de estilo ─────────────────────────────────────────────────────────

def _thin_border():
    s = Side(style='thin', color=BORDER_C)
    return Border(left=s, right=s, top=s, bottom=s)

def _h(ws, row, col, value, bold=True, color=HDR_FG, bg=HDR_BG, size=10,
        halign='center', wrap=True):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=size)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical='center', wrap_text=wrap)
    c.border = _thin_border()
    return c

def _d(ws, row, col, value, bg=WHITE, bold=False, halign='left',
        number_format=None, color='000000'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, size=10)
    c.fill = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal=halign, vertical='center')
    c.border = _thin_border()
    if number_format:
        c.number_format = number_format
    return c

def _col(n):
    return get_column_letter(n)

def _freeze(ws, cell):
    ws.freeze_panes = cell

def _autofilter(ws, ref):
    ws.auto_filter.ref = ref

def _width(ws, mapping):
    for col, w in mapping.items():
        ws.column_dimensions[_col(col)].width = w

def _row_height(ws, row, h):
    ws.row_dimensions[row].height = h

def _title_row(ws, row, ncols, text, bg=TITLE_BG):
    ws.merge_cells(start_row=row, start_column=1,
                   end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, color=TITLE_FG, size=12)
    c.fill      = PatternFill('solid', fgColor=bg)
    c.alignment = Alignment(horizontal='left', vertical='center')
    _row_height(ws, row, 28)


# ── Sheet 1: Mensal ───────────────────────────────────────────────────────────

MENSAL_COLS = {
    1: ('Mês',              8),
    2: ('Ano',              7),
    3: ('Período',         10),
    4: ('Valor RV\n(R$)',  14),
    5: ('Investido RV\n(R$)', 15),
    6: ('Ganho/Perda\n(R$)', 14),
    7: ('SELIC\nMensal (R$)', 14),
    8: ('SELIC\nAcumulado (R$)', 16),
    9: ('Div. FIIs\n(R$)', 13),
    10:('Div. Ações\n(R$)', 13),
    11:('Total\nRenda (R$)', 13),
    12:('Aporte\n(R$)',     12),
    13:('META\n(%)',         9),
    14:('REAL\n(%)',         9),
    15:('Observação',       22),
}

def _create_mensal(wb, monthly):
    ws = wb.create_sheet('Mensal')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '1e3a5f'

    _title_row(ws, 1, len(MENSAL_COLS), '📅  HISTÓRICO MENSAL — CARTEIRA VARIÁVEL')

    # Header
    for col, (label, width) in MENSAL_COLS.items():
        _h(ws, 2, col, label)
        ws.column_dimensions[_col(col)].width = width
    _row_height(ws, 2, 36)
    _freeze(ws, 'A3')
    _autofilter(ws, f'A2:{_col(len(MENSAL_COLS))}2')

    # Data
    for i, m in enumerate(monthly):
        r = i + 3
        bg = ALT_BG if i % 2 == 0 else WHITE

        mn   = m['month']
        yr   = m['year']
        pv   = m['portfolio_value']
        inv  = m['total_invested_var']
        gl   = m['gain_loss']
        sel  = m['selic']
        selc = m['selic_cumulative']
        fii  = m['fii_dividends']
        stk  = m['stock_dividends']
        appr = m['appreciation'] or 0
        meta = m['meta']
        real = m['real']

        total_renda = (sel or 0) + (fii or 0) + (stk or 0) + appr

        _d(ws, r, 1,  mn.capitalize(),              bg, halign='center')
        _d(ws, r, 2,  yr,                            bg, halign='center')
        _d(ws, r, 3,  f"{mn.capitalize()}/{str(yr)[2:]}", bg, halign='center', bold=True)
        _d(ws, r, 4,  pv,   bg, halign='right', number_format='R$ #,##0.00')
        _d(ws, r, 5,  inv,  bg, halign='right', number_format='R$ #,##0.00')

        # Ganho/perda — cor condicional manual
        gl_bg = GREEN_BG if (gl or 0) >= 0 else RED_BG
        _d(ws, r, 6,  gl,   gl_bg, halign='right', number_format='R$ #,##0.00')

        _d(ws, r, 7,  sel,  bg, halign='right', number_format='R$ #,##0.00')
        _d(ws, r, 8,  selc, bg, halign='right', number_format='R$ #,##0.00')
        _d(ws, r, 9,  fii,  bg, halign='right', number_format='R$ #,##0.00')
        _d(ws, r, 10, stk,  bg, halign='right', number_format='R$ #,##0.00')
        _d(ws, r, 11, total_renda, BLUE_BG, halign='right', bold=True,
           number_format='R$ #,##0.00')
        _d(ws, r, 12, None, bg, halign='right', number_format='R$ #,##0.00')

        # META/REAL — cor condicional
        _d(ws, r, 13, meta, bg, halign='center', number_format='0.00%')
        real_bg = GREEN_BG if real and meta and real >= meta else (RED_BG if real else bg)
        _d(ws, r, 14, real, real_bg, halign='center', number_format='0.00%')
        _d(ws, r, 15, None, bg)

    # Linha de totais
    last_r = len(monthly) + 2
    total_r = last_r + 1
    ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=3)
    _h(ws, total_r, 1, 'TOTAIS', bg='0f172a')
    for col in range(4, 14):
        col_letter = _col(col)
        formula = f'=SUM({col_letter}3:{col_letter}{last_r})'
        c = ws.cell(row=total_r, column=col, value=formula)
        c.font   = Font(bold=True, color=HDR_FG, size=10)
        c.fill   = PatternFill('solid', fgColor='0f172a')
        c.border = _thin_border()
        c.alignment = Alignment(horizontal='right', vertical='center')
        if col in (13, 14):
            c.number_format = '0.00%'
            c.value = None
        else:
            c.number_format = 'R$ #,##0.00'

    # Nota de rodapé
    note_r = total_r + 2
    ws.merge_cells(start_row=note_r, start_column=1, end_row=note_r, end_column=15)
    c = ws.cell(row=note_r, column=1,
                value='ℹ️  Ganho/Perda = Valor RV atual − Total Investido em RV  |  Total Renda = SELIC + FIIs + Ações + Valorização')
    c.font = Font(italic=True, color='64748b', size=9)
    c.alignment = Alignment(horizontal='left')


# ── Sheet 2: Dividendos ───────────────────────────────────────────────────────

DIV_COLS = {
    1: ('Mês',          8),
    2: ('Ano',          7),
    3: ('Período',     10),
    4: ('Data Pgto',   12),
    5: ('Ativo',       10),
    6: ('Tipo',        10),
    7: ('Valor (R$)',  12),
    8: ('Observação',  28),
}

TYPE_COLORS = {
    'SELIC':       'dbeafe',
    'FII':         'd1fae5',
    'Ação':        'fef9c3',
    'Valorização': 'ede9fe',
    'ETF':         'e0e7ff',
    'Cripto':      'fce7f3',
}

def _create_dividendos(wb, monthly):
    ws = wb.create_sheet('Dividendos')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '059669'

    _title_row(ws, 1, len(DIV_COLS), '💰  HISTÓRICO DE DIVIDENDOS E RENDA MENSAL',
               bg='064e3b')

    for col, (label, width) in DIV_COLS.items():
        _h(ws, 2, col, label, bg='059669')
        ws.column_dimensions[_col(col)].width = width
    _row_height(ws, 2, 32)

    # Dropdown para Tipo
    dv = DataValidation(
        type='list',
        formula1='"SELIC,FII,Ação,Valorização,ETF,Cripto"',
        showDropDown=False,
        showErrorMessage=True,
        errorTitle='Tipo inválido',
        error='Escolha: SELIC, FII, Ação, Valorização, ETF ou Cripto'
    )
    ws.add_data_validation(dv)

    _freeze(ws, 'A3')
    _autofilter(ws, f'A2:{_col(len(DIV_COLS))}2')

    row = 3
    for m in monthly:
        mn  = m['month']
        yr  = m['year']
        lbl = f"{mn.capitalize()}/{str(yr)[2:]}"

        # SELIC
        if m['selic']:
            bg = TYPE_COLORS['SELIC']
            _d(ws, row, 1, mn.capitalize(), bg, halign='center')
            _d(ws, row, 2, yr,             bg, halign='center')
            _d(ws, row, 3, lbl,            bg, halign='center', bold=True)
            _d(ws, row, 4, None,           bg, halign='center')
            _d(ws, row, 5, 'SELIC',        bg, halign='center', bold=True)
            _d(ws, row, 6, 'SELIC',        bg, halign='center')
            _d(ws, row, 7, m['selic'],     bg, halign='right', number_format='R$ #,##0.00')
            _d(ws, row, 8, None,           bg)
            dv.add(ws.cell(row=row, column=6))
            row += 1

        # Ativos (FII + Ações)
        for asset in m['assets']:
            tipo = 'FII' if _is_fii(asset['sku']) else 'Ação'
            bg = TYPE_COLORS.get(tipo, WHITE)
            dt = asset['date']
            if dt:
                try:
                    dt = datetime.strptime(dt, '%Y-%m-%d').date()
                except ValueError:
                    dt = None

            _d(ws, row, 1, mn.capitalize(), bg, halign='center')
            _d(ws, row, 2, yr,             bg, halign='center')
            _d(ws, row, 3, lbl,            bg, halign='center', bold=True)
            c_date = _d(ws, row, 4, dt,   bg, halign='center')
            if dt:
                c_date.number_format = 'DD/MM/YYYY'
            _d(ws, row, 5, asset['sku'].upper(), bg, halign='center', bold=True)
            _d(ws, row, 6, tipo,               bg, halign='center')
            _d(ws, row, 7, asset['value'],      bg, halign='right',
               number_format='R$ #,##0.00')
            _d(ws, row, 8, None, bg)
            dv.add(ws.cell(row=row, column=6))
            row += 1

        # Valorização de ações de crescimento
        if m.get('appreciation'):
            bg = TYPE_COLORS['Valorização']
            _d(ws, row, 1, mn.capitalize(),      bg, halign='center')
            _d(ws, row, 2, yr,                   bg, halign='center')
            _d(ws, row, 3, lbl,                  bg, halign='center', bold=True)
            _d(ws, row, 4, None,                 bg)
            _d(ws, row, 5, 'Valorização',        bg, halign='center', bold=True)
            _d(ws, row, 6, 'Valorização',        bg, halign='center')
            _d(ws, row, 7, m['appreciation'],    bg, halign='right',
               number_format='R$ #,##0.00')
            _d(ws, row, 8, 'Proporcional ao tempo de posse', bg)
            dv.add(ws.cell(row=row, column=6))
            row += 1

    # Linha de total
    total_r = row
    ws.merge_cells(start_row=total_r, start_column=1,
                   end_row=total_r, end_column=6)
    _h(ws, total_r, 1, 'TOTAL ACUMULADO', bg='064e3b')
    c = ws.cell(row=total_r, column=7,
                value=f'=SUM(G3:G{row - 1})')
    c.font   = Font(bold=True, color=HDR_FG, size=10)
    c.fill   = PatternFill('solid', fgColor='064e3b')
    c.border = _thin_border()
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = 'R$ #,##0.00'


# ── Sheet 3: Carteira Atual ───────────────────────────────────────────────────

CART_COLS = {
    1: ('Ativo',           12),
    2: ('Ticker Yahoo',    14),
    3: ('Tipo',            10),
    4: ('Valor Atual\n(R$)', 15),
    5: ('% Carteira\nVariável', 14),
    6: ('Qtd. Cotas\n(aprox.)', 14),
    7: ('Preço Médio\n(R$)', 14),
    8: ('Última\nAtualização', 16),
    9: ('Observação',      25),
}

YAHOO_MAP = {
    'petr4':  'PETR4.SA',  'roxo34': 'ROXO34.SA', 'hapv3':  'HAPV3.SA',
    'aure3f': 'AURE3.SA',  'b3sa3':  'B3SA3.SA',  'bbas3':  'BBAS3.SA',
    'recr11': 'RECR11.SA', 'simh3':  'SIMH3.SA',  'tgar11': 'TGAR11.SA',
    'trxf11': 'TRXF11.SA',
}
TYPE_MAP = {
    'petr4':'Ação','roxo34':'ETF','hapv3':'Ação','aure3f':'Ação',
    'b3sa3':'Ação','bbas3':'Ação','recr11':'FII','simh3':'Ação',
    'tgar11':'FII','trxf11':'FII',
}

def _create_carteira(wb, summary):
    ws = wb.create_sheet('Carteira_Atual')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = '7c3aed'

    _title_row(ws, 1, len(CART_COLS),
               '📊  CARTEIRA VARIÁVEL — POSIÇÕES ATUAIS', bg='3b0764')

    for col, (label, width) in CART_COLS.items():
        _h(ws, 2, col, label, bg='7c3aed')
        ws.column_dimensions[_col(col)].width = width
    _row_height(ws, 2, 36)
    _freeze(ws, 'A3')

    # Dropdown Tipo
    dv = DataValidation(
        type='list',
        formula1='"Ação,FII,ETF,Cripto"',
        showDropDown=False
    )
    ws.add_data_validation(dv)

    alloc = summary.get('asset_allocation', [])
    total_var = sum(a['value'] for a in alloc) or 1

    for i, a in enumerate(sorted(alloc, key=lambda x: x['value'], reverse=True)):
        r   = i + 3
        bg  = ALT_BG if i % 2 == 0 else WHITE
        sku = a['sku']
        tipo = TYPE_MAP.get(sku, 'Ação')
        ticker = YAHOO_MAP.get(sku, sku.upper() + '.SA')
        pct = a['value'] / total_var

        _d(ws, r, 1, sku.upper(),    bg, bold=True, halign='center')
        _d(ws, r, 2, ticker,         bg, halign='center', color='2563eb')
        tipo_c = _d(ws, r, 3, tipo,  bg, halign='center')
        dv.add(tipo_c)
        _d(ws, r, 4, a['value'],     bg, halign='right', number_format='R$ #,##0.00')
        _d(ws, r, 5, pct,            bg, halign='center', number_format='0.0%')
        _d(ws, r, 6, None,           bg, halign='right')
        _d(ws, r, 7, None,           bg, halign='right', number_format='R$ #,##0.00')
        today = date.today()
        c_date = _d(ws, r, 8, today, bg, halign='center')
        c_date.number_format = 'DD/MM/YYYY'
        _d(ws, r, 9, None,           bg)

    # Total
    last_r = len(alloc) + 2
    total_r = last_r + 1
    ws.merge_cells(start_row=total_r, start_column=1, end_row=total_r, end_column=3)
    _h(ws, total_r, 1, 'TOTAL VARIÁVEL', bg='3b0764')
    c = ws.cell(row=total_r, column=4, value=f'=SUM(D3:D{last_r})')
    c.font = Font(bold=True, color=HDR_FG, size=10)
    c.fill = PatternFill('solid', fgColor='3b0764')
    c.border = _thin_border()
    c.alignment = Alignment(horizontal='right', vertical='center')
    c.number_format = 'R$ #,##0.00'


# ── Sheet 4: Resumo ───────────────────────────────────────────────────────────

def _create_resumo(wb, summary):
    ws = wb.create_sheet('Resumo')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = 'dc2626'

    _title_row(ws, 1, 4, '💼  RESUMO GERAL DA CARTEIRA', bg='450a0a')

    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 28

    # Cabeçalho das colunas
    for col, txt in [(1,'Categoria'), (2,'Valor Atual (R$)'), (3,'Meta'), (4,'Observação')]:
        _h(ws, 2, col, txt, bg='b91c1c')
    _row_height(ws, 2, 28)

    s = summary

    rows_data = [
        # (Seção, label, valor, meta, obs)
        ('ALOCAÇÃO', 'Renda Variável (R.V.)',       s.get('variable'),             None,       'Ações + FIIs + ETFs'),
        ('ALOCAÇÃO', 'Renda Fixa — SELIC',          s.get('fixed'),                None,       ''),
        ('ALOCAÇÃO', 'Cripto',                       s.get('crypto'),               None,       'Meta: 10% da carteira'),
        ('ALOCAÇÃO', 'R.E. — Reserva Emergência',   s.get('reserve_emergency'),    'R$ 4.000', 'Meta ATINGIDA ✓'),
        ('ALOCAÇÃO', 'R.O. — Reserva Oportunidade', s.get('reserve_opportunity'),  'R$ 2.000', 'Falta para completar'),
        ('ALOCAÇÃO', 'Stand By (troco)',             s.get('stand_by'),             None,       'Sobra não investida'),
        (None, None, None, None, None),
        ('RESULTADO', 'Total Investido (sem R.E./R.O.)', s.get('total_invested'),  None, ''),
        ('RESULTADO', 'Total Inicial (Jun/2024)',        s.get('total_initial'),   None, ''),
        ('RESULTADO', 'Total Ideal (META acumulada)',    s.get('total_ideal'),     None, ''),
        ('RESULTADO', 'Total Real (valor atual total)',  s.get('total_real'),      None, ''),
        (None, None, None, None, None),
        ('RENDIMENTO', 'Total Aportes',              s.get('total_contributions'),  None, ''),
        ('RENDIMENTO', 'Lucro Total Acumulado',      s.get('total_profit'),         None, ''),
        ('RENDIMENTO', '  → SELIC',                  s.get('selic_profit'),         None, ''),
        ('RENDIMENTO', '  → FIIs (dividendos)',      s.get('total_fii_dividends'),  None, ''),
        ('RENDIMENTO', '  → Ações (dividendos)',     s.get('total_stock_dividends'),None, ''),
        ('RENDIMENTO', 'Cripto P&L',                 s.get('crypto_pnl'),           None, 'Negativo = prejuízo acumulado'),
    ]

    SECTION_COLORS = {
        'ALOCAÇÃO':  ('1d4ed8', 'dbeafe'),
        'RESULTADO': ('065f46', 'd1fae5'),
        'RENDIMENTO':('7c2d12', 'ffedd5'),
    }

    current_section = None
    r = 3
    for sec, label, value, meta, obs in rows_data:
        if label is None:
            r += 1
            continue

        if sec != current_section:
            current_section = sec
            hdr_dark, _ = SECTION_COLORS[sec]
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
            c = ws.cell(row=r, column=1, value=f'▸  {sec}')
            c.font      = Font(bold=True, color='FFFFFF', size=10)
            c.fill      = PatternFill('solid', fgColor=hdr_dark)
            c.alignment = Alignment(horizontal='left', vertical='center',
                                    indent=1)
            c.border    = _thin_border()
            _row_height(ws, r, 22)
            r += 1

        _, row_bg = SECTION_COLORS[sec]
        _d(ws, r, 1, label, row_bg, halign='left')
        c_val = _d(ws, r, 2, value, row_bg, halign='right', bold=True,
                   number_format='R$ #,##0.00')
        if value and value < 0:
            c_val.font = Font(bold=True, color='dc2626', size=10)
        _d(ws, r, 3, meta, row_bg, halign='center')
        _d(ws, r, 4, obs,  row_bg, halign='left', color='475569')
        _row_height(ws, r, 20)
        r += 1

    # Última atualização
    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    c = ws.cell(row=r, column=1,
                value=f'Última atualização: {date.today().strftime("%d/%m/%Y")}')
    c.font = Font(italic=True, color='94a3b8', size=9)
    c.alignment = Alignment(horizontal='right')


# ── Sheet 5: Como Usar ────────────────────────────────────────────────────────

def _create_como_usar(wb):
    ws = wb.create_sheet('Como Usar')
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = 'f59e0b'
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 70

    _title_row(ws, 1, 3, '📖  COMO USAR ESTA PLANILHA', bg='78350f')
    _row_height(ws, 1, 30)

    instrucoes = [
        ('📅 ABA MENSAL', None, 'azul escuro'),
        ('Quando preencher?', 'Uma vez por mes, geralmente no ultimo dia util do mes.',         None),
        ('Valor RV',         'O valor total de mercado das suas acoes + FIIs naquele momento.', None),
        ('Investido RV',     'Quanto voce ja investiu em renda variavel no total (historico).', None),
        ('Ganho/Perda',      'Calculado automaticamente: Valor RV - Investido RV. Negativo = abaixo do custo.', None),
        ('SELIC Mensal',     'Rendimento da SELIC naquele mes (juros recebidos).',               None),
        ('SELIC Acumulado',  'Soma de todos os rendimentos SELIC desde o inicio.',               None),
        ('Div. FIIs',        'Soma dos dividendos recebidos de FIIs no mes.',                    None),
        ('Div. Acoes',       'Soma dos dividendos de acoes no mes.',                             None),
        ('Total Renda',      'SELIC + FIIs + Acoes + Valorizacao — renda total do mes.',        None),
        ('Aporte',           'Quanto voce investiu de novo naquele mes.',                        None),
        ('META',             'Meta de rendimento mensal — padrao 0,95%. Nao precisa alterar.',  None),
        ('REAL',             'Rendimento real do mes em %. Ex: 0,87% — escreva 0,0087.',        None),
        (None, None, None),
        ('ABA DIVIDENDOS', None, 'verde'),
        ('Quando preencher?', 'Cada vez que receber um dividendo, rendimento ou pagamento.',     None),
        ('Ativo',            'Nome do ativo que pagou. Ex: RECR11, BBAS3, SELIC.',              None),
        ('Tipo',             'Use o dropdown: FII / Acao / SELIC / Valorizacao / ETF / Cripto.',None),
        ('Data Pgto',        'Data em que caiu na conta. Formato DD/MM/AAAA.',                  None),
        ('Valor (R$)',       'Valor bruto recebido em reais.',                                   None),
        (None, None, None),
        ('ABA CARTEIRA_ATUAL', None, 'roxo'),
        ('Quando preencher?', 'Uma vez por mes ou apos cada compra/venda.',                     None),
        ('Valor Atual (R$)', 'Valor de mercado atual da sua posicao (preco x quantidade).',     None),
        ('Qtd. Cotas',       'Quantidade de cotas/acoes que voce possui.',                      None),
        ('Preco Medio',      'Seu preco medio de compra (opcional, para controle de custo).',   None),
        (None, None, None),
        ('ABA RESUMO', None, 'vermelho'),
        ('Quando preencher?', 'Uma vez por mes, atualize os valores de cada categoria.',        None),
        ('Renda Variavel',   'Valor atual de mercado de todas as acoes + FIIs juntos.',         None),
        ('SELIC',            'Saldo atual em renda fixa (SELIC).',                              None),
        ('Cripto',           'Valor atual da carteira de criptomoedas.',                        None),
        ('R.E.',             'Reserva de Emergencia (meta: R$ 4.000).',                         None),
        ('R.O.',             'Reserva de Oportunidade (meta: R$ 2.000 — para aproveitar quedas).', None),
        (None, None, None),
        ('IMPORTANTE', None, 'amarelo'),
        ('Stand By',         'NAO tem meta de alocacao — e apenas o troco que sobrou no mes.',  None),
        ('R.O. x Stand By',  'R.O. e intencional (guardado propositalmente). Stand by e acidental.', None),
        ('Valorizacao',      'Acoes de crescimento (HAPV3, SIMH3) — valor proporcional ao tempo de posse.', None),
        ('App Web',          'Apos preencher, salve o arquivo e recarregue o dashboard no navegador.', None),
    ]

    SECTION_BG_MAP = {
        'azul escuro': ('1e3a5f', 'dbeafe'),
        'verde':       ('065f46', 'd1fae5'),
        'roxo':        ('3b0764', 'ede9fe'),
        'vermelho':    ('7f1d1d', 'fee2e2'),
        'amarelo':     ('78350f', 'fef9c3'),
    }
    current_col = 'dbeafe'
    r = 3
    for label, desc, extra in instrucoes:
        if label is None:
            r += 1
            continue

        if desc is None and extra in SECTION_BG_MAP:
            # Section header
            dark, light = SECTION_BG_MAP[extra]
            current_col = light
            ws.merge_cells(start_row=r, start_column=2,
                           end_row=r, end_column=3)
            c = ws.cell(row=r, column=2, value=label)
            c.font      = Font(bold=True, color='FFFFFF', size=11)
            c.fill      = PatternFill('solid', fgColor=dark)
            c.alignment = Alignment(horizontal='left', vertical='center',
                                    indent=1)
            c.border    = _thin_border()
            _row_height(ws, r, 24)
        else:
            c1 = ws.cell(row=r, column=2, value=label)
            c1.font      = Font(bold=True, color='1e293b', size=10)
            c1.fill      = PatternFill('solid', fgColor=current_col)
            c1.alignment = Alignment(horizontal='left', vertical='center',
                                     indent=1)
            c1.border    = _thin_border()
            c2 = ws.cell(row=r, column=3, value=desc)
            c2.font      = Font(color='334155', size=10)
            c2.fill      = PatternFill('solid', fgColor='f8fafc')
            c2.alignment = Alignment(horizontal='left', vertical='center',
                                     wrap_text=True, indent=1)
            c2.border    = _thin_border()
            _row_height(ws, r, 18)
        r += 1


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print('Lendo dados de finance.xlsx...')
    data = parse_finance_data('finance.xlsx')
    monthly = data['monthly']
    summary = data['summary']
    print(f'  {len(monthly)} meses encontrados ({monthly[0]["label"]} a {monthly[-1]["label"]})')

    print('Criando finance_data.xlsx...')
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    _create_mensal(wb, monthly)
    print('  OK Aba Mensal')

    _create_dividendos(wb, monthly)
    print('  ✓ Aba Dividendos')

    _create_carteira(wb, summary)
    print('  ✓ Aba Carteira_Atual')

    _create_resumo(wb, summary)
    print('  ✓ Aba Resumo')

    _create_como_usar(wb)
    print('  ✓ Aba Como Usar')

    wb.save('finance_data.xlsx')
    print('\nSUCESSO: finance_data.xlsx criado com sucesso!')
    print('   Abra o arquivo para verificar e comece a alimentar os dados mensalmente.')


if __name__ == '__main__':
    main()
