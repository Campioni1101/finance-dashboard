"""
Parser que lê finance_data.xlsx (novo formato organizado).
Fallback automático para finance.xlsx (formato legado) se o novo não existir.
"""
import os
import unicodedata
import openpyxl
from datetime import datetime


def _norm(s: str) -> str:
    """Remove acentos, converte para minúsculas e normaliza espaços/setas/travessões."""
    s = s.replace('→', '->').replace('—', '-').replace('–', '-')
    s = unicodedata.normalize('NFD', s).encode('ascii', 'ignore').decode()
    return s.lower().strip()

CALENDAR_ORDER = {
    'jan': 1, 'fev': 2, 'mar': 3, 'abr': 4, 'mai': 5, 'jun': 6,
    'jul': 7, 'ago': 8, 'set': 9, 'out': 10, 'nov': 11, 'dez': 12
}

FII_PATTERNS = ('11', 'tgar', 'recr', 'trxf')
STOCK_SKUS = {
    'bbas3', 'b3sa3', 'bbsa3', 'bbasa3', 'aure3', 'aure3f',
    'simh3', 'hapv3', 'petr4', 'roxo34'
}


def _is_fii(sku: str) -> bool:
    s = sku.lower().strip()
    return s.endswith('11') or any(s.startswith(p) or p in s for p in FII_PATTERNS)


def _is_stock(sku: str) -> bool:
    s = sku.lower().strip()
    return any(sk in s for sk in STOCK_SKUS)


def parse_finance_data(filepath: str) -> dict:
    """
    Detecta automaticamente o formato do arquivo:
    - finance_data.xlsx  → novo formato (abas: Mensal, Dividendos, Carteira_Atual, Resumo)
    - finance.xlsx       → formato legado (abas: True, Planilha3)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets = wb.sheetnames

    if 'Mensal' in sheets:
        return _parse_new_format(wb)
    return _parse_legacy_format(wb)


# ── Novo formato (finance_data.xlsx) ─────────────────────────────────────────

def _parse_new_format(wb) -> dict:
    monthly = _parse_mensal_sheet(wb['Mensal'])
    summary = _parse_resumo_sheet(wb)
    return {'monthly': monthly, 'summary': summary}


def _parse_mensal_sheet(ws) -> list:
    """Lê a aba Mensal: uma linha por mês, colunas fixas."""
    # Colunas: Mês(1) Ano(2) Período(3) ValorRV(4) InvestidoRV(5) GanhoPert(6)
    #          SelicMensal(7) SelicAcum(8) DivFII(9) DivAcoes(10) TotalRenda(11)
    #          Aporte(12) META(13) REAL(14) Obs(15)
    blocks = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        mn = row[0]
        if not isinstance(mn, str) or mn.strip().lower() not in CALENDAR_ORDER:
            continue
        mn = mn.strip().lower()
        yr = row[1] if isinstance(row[1], int) else None
        if not yr:
            continue

        def v(i):
            val = row[i] if len(row) > i else None
            return val if isinstance(val, (int, float)) else None

        pv  = v(3)
        inv = v(4)
        gl  = v(5)
        sel = v(6)
        selc= v(7)
        fii = v(8)
        stk = v(9)
        tot = v(10)
        meta= v(12)
        real= v(13)

        blocks.append({
            'month':           mn,
            'year':            yr,
            'label':           f"{mn.capitalize()}/{str(yr)[2:]}",
            'portfolio_value': pv,
            'gain_loss':       gl if gl is not None else ((pv - inv) if pv and inv else None),
            'total_invested_var': inv,
            'selic':           sel,
            'selic_cumulative':selc,
            'fii_dividends':   round(fii, 2) if fii else 0.0,
            'stock_dividends': round(stk, 2) if stk else 0.0,
            'appreciation':    None,
            'total_dividends': tot,
            'meta':            meta,
            'real':            real,
            'assets':          [],  # detalhe por ativo vem da aba Dividendos
        })

    # Preencher assets a partir da aba Dividendos se disponível
    return blocks


def _enrich_with_dividends(wb, blocks: list):
    """Popula blocks[].assets a partir da aba Dividendos."""
    if 'Dividendos' not in wb.sheetnames:
        return
    ws = wb['Dividendos']
    # Colunas: Mês(1) Ano(2) Período(3) DataPgto(4) Ativo(5) Tipo(6) Valor(7) Obs(8)
    lookup = {(b['month'], b['year']): b for b in blocks}
    for row in ws.iter_rows(min_row=3, values_only=True):
        mn  = row[0]
        yr  = row[1]
        sku = row[4]
        typ = row[5]
        val = row[6]
        dt  = row[3]

        if not (isinstance(mn, str) and isinstance(yr, int)
                and isinstance(sku, str) and isinstance(val, (int, float))):
            continue
        if typ and typ.upper() in ('SELIC', 'VALORIZACAO', 'VALORIZACAO'):
            continue  # já capturado nas colunas agregadas

        key = (mn.strip().lower(), yr)
        block = lookup.get(key)
        if not block:
            continue

        date_str = None
        if isinstance(dt, (datetime,)):
            date_str = dt.strftime('%Y-%m-%d')
        elif isinstance(dt, str):
            date_str = dt

        block['assets'].append({
            'sku':   sku.strip(),
            'value': round(val, 2),
            'date':  date_str,
        })


def _parse_resumo_sheet(wb) -> dict:
    """Lê as abas Resumo e Carteira_Atual para montar o summary."""
    s: dict = {
        'variable': None, 'fixed': None, 'crypto': None, 'stand_by': None,
        'reserve_emergency': 4000.0, 'reserve_opportunity': 1671.55,
        'total_profit': None, 'selic_profit': None, 'selic_profit_pct': None,
        'total_contributions': None, 'crypto_pnl': None,
        'fii_dividends': {}, 'total_fii_dividends': None,
        'stock_dividends': {}, 'total_stock_dividends': None,
        'total_invested': None, 'total_initial': None,
        'total_ideal': None, 'total_real': None,
        'asset_allocation': [],
    }

    if 'Resumo' not in wb.sheetnames:
        return s

    ws_r = wb['Resumo']
    LABEL_MAP = {
        'renda variavel':                    'variable',
        'renda variavel (r.v.)':             'variable',
        'renda fixa - selic':                'fixed',
        'selic':                             'fixed',
        'cripto':                            'crypto',
        'r.e. - reserva emergencia':         're',
        'r.o. - reserva oportunidade':       'ro',
        'stand by (troco)':                  'stand_by',
        'total investido (sem r.e./r.o.)':   'total_invested',
        'total inicial (jun/2024)':          'total_initial',
        'total ideal (meta acumulada)':      'total_ideal',
        'total real (valor atual total)':    'total_real',
        'total aportes':                     'total_contributions',
        'lucro total acumulado':             'total_profit',
        '-> selic':                          'selic_profit',
        '-> fiis (dividendos)':              'total_fii_dividends',
        '-> acoes (dividendos)':             'total_stock_dividends',
        'cripto p&l':                        'crypto_pnl',
    }
    for row in ws_r.iter_rows(min_row=3, values_only=True):
        label = row[0]
        value = row[1]
        if not isinstance(label, str) or not isinstance(value, (int, float)):
            continue
        key = LABEL_MAP.get(_norm(label))
        if key == 're':
            s['reserve_emergency'] = value
        elif key == 'ro':
            s['reserve_opportunity'] = value
        elif key:
            s[key] = value

    # Carteira_Atual → asset_allocation
    if 'Carteira_Atual' in wb.sheetnames:
        ws_c = wb['Carteira_Atual']
        alloc = []
        for row in ws_c.iter_rows(min_row=3, values_only=True):
            sku = row[0]
            val = row[3]
            pct = row[4]
            if isinstance(sku, str) and isinstance(val, (int, float)) and val > 0:
                alloc.append({
                    'sku':        sku.lower(),
                    'value':      round(val, 2),
                    'percentage': round(pct * 100, 2) if isinstance(pct, float) else None,
                })
        s['asset_allocation'] = alloc

    # Dividendos por ativo (se aba existir)
    if 'Dividendos' in wb.sheetnames:
        ws_d = wb['Dividendos']
        fii_acc: dict = {}
        stk_acc: dict = {}
        for row in ws_d.iter_rows(min_row=3, values_only=True):
            sku = row[4]
            typ = row[5]
            val = row[6]
            if not (isinstance(sku, str) and isinstance(val, (int, float)) and val > 0):
                continue
            sku_n = sku.strip().upper()
            typ_n = (typ or '').strip().upper()
            if typ_n == 'FII':
                fii_acc[sku_n] = round(fii_acc.get(sku_n, 0) + val, 2)
            elif typ_n in ('ACAO', 'AÇÃO', 'ACÃO', 'ACAO'):
                stk_acc[sku_n] = round(stk_acc.get(sku_n, 0) + val, 2)
        if fii_acc:
            s['fii_dividends'] = {k.lower(): v for k, v in fii_acc.items()}
        if stk_acc:
            s['stock_dividends'] = {k.lower(): v for k, v in stk_acc.items()}

    return s


# ── Formato legado (finance.xlsx) ─────────────────────────────────────────────

def _parse_legacy_format(wb) -> dict:
    monthly = _parse_legacy_monthly(wb['True'])
    summary = _parse_legacy_summary(wb['Planilha3'])
    return {'monthly': monthly, 'summary': summary}


def _parse_legacy_monthly(ws) -> list:
    rows = list(ws.iter_rows(values_only=True))
    blocks = []
    current = None
    year = 2024
    prev_cal = 0

    for row in rows:
        a = row[0] if row else None
        b = row[1] if len(row) > 1 else None
        c = row[2] if len(row) > 2 else None

        if isinstance(a, str) and a.strip().lower() in CALENDAR_ORDER:
            mn = a.strip().lower()
            cal = CALENDAR_ORDER[mn]
            if prev_cal > 0 and cal < prev_cal:
                year += 1
            prev_cal = cal

            if current:
                _finalize_legacy_block(current)
                blocks.append(current)

            pv = b if isinstance(b, (int, float)) else None
            gl = c if isinstance(c, (int, float)) else None

            current = {
                'month': mn, 'year': year,
                'label': f"{mn.capitalize()}/{str(year)[2:]}",
                'portfolio_value': pv,
                'gain_loss': gl,
                'total_invested_var': (pv - gl) if (pv is not None and gl is not None) else None,
                'selic': None, 'selic_cumulative': None,
                'assets': [], 'appreciation': None,
                'total_dividends': None, 'meta': None, 'real': None,
            }

        elif current is not None and isinstance(a, str):
            al = a.strip().lower()
            if al == 'sku':
                continue
            elif 'selic' in al:
                if isinstance(b, (int, float)):
                    current['selic'] = b
                if isinstance(c, (int, float)):
                    current['selic_cumulative'] = c
            elif al == 'total':
                if isinstance(b, (int, float)):
                    current['total_dividends'] = b
            elif al == 'meta':
                if isinstance(b, (int, float)):
                    current['meta'] = b
            elif al == 'real':
                if isinstance(b, (int, float)):
                    current['real'] = b
            elif al == 'valor':
                if isinstance(b, (int, float)) and b > 0:
                    current['appreciation'] = round(b, 2)
            elif isinstance(b, (int, float)) and b > 0:
                date_val = row[2] if len(row) > 2 else None
                date_str = date_val.strftime('%Y-%m-%d') if isinstance(date_val, datetime) else None
                current['assets'].append({
                    'sku': a.strip(), 'value': round(b, 2), 'date': date_str
                })

    if current:
        _finalize_legacy_block(current)
        blocks.append(current)

    return blocks


def _finalize_legacy_block(block: dict):
    fii_total = 0.0
    stock_total = 0.0
    for asset in block['assets']:
        if _is_fii(asset['sku']):
            fii_total += asset['value']
        else:
            stock_total += asset['value']
    block['fii_dividends'] = round(fii_total, 2)
    block['stock_dividends'] = round(stock_total, 2)


def _parse_legacy_summary(ws) -> dict:
    rows = list(ws.iter_rows(values_only=True))

    def val(r, i):
        return r[i] if len(r) > i and isinstance(r[i], (int, float)) else None

    def row(idx):
        return rows[idx] if len(rows) > idx else []

    r2, r3, r4, r5, r6, r7, r8 = (row(i) for i in range(1, 8))
    r10, r12, r14, r16 = row(9), row(11), row(13), row(15)

    asset_allocation = []
    for i in range(17, min(28, len(rows))):
        r = rows[i]
        v = r[0] if len(r) > 0 and isinstance(r[0], (int, float)) else None
        n = r[1] if len(r) > 1 and isinstance(r[1], str) else None
        p = r[2] if len(r) > 2 and isinstance(r[2], float) else None
        if v is not None and n:
            asset_allocation.append({
                'sku': n.lower(),
                'value': round(v, 2),
                'percentage': round(p * 100, 2) if p else None
            })

    return {
        'variable': val(r2, 0), 'fixed': val(r4, 0),
        'crypto': val(r6, 0),   'stand_by': val(r8, 0),
        'reserve_emergency': 4000.0, 'reserve_opportunity': 1671.55,
        'total_profit': val(r2, 2), 'selic_profit': val(r2, 3),
        'selic_profit_pct': val(r3, 3), 'total_contributions': val(r2, 10),
        'crypto_pnl': val(r2, 9),
        'fii_dividends': {
            'tgar11': val(r2, 5), 'recr11': val(r3, 5), 'trxf11': val(r4, 5),
        },
        'total_fii_dividends': val(r5, 5),
        'stock_dividends': {
            'bbas3': val(r2, 7), 'b3sa3': val(r3, 7), 'aure3': val(r4, 7),
            'simh3': val(r5, 7), 'hapv3': val(r6, 7), 'petr4': val(r7, 7),
        },
        'total_stock_dividends': val(r8, 7),
        'total_invested': val(r10, 0), 'total_initial': val(r12, 0),
        'total_ideal': val(r14, 0),    'total_real': val(r16, 0),
        'asset_allocation': asset_allocation,
    }
